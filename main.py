import sys
import datetime
import numpy as np
from PyQt5 import uic
from PyQt5.QtWidgets import QApplication, QMainWindow, QDateTimeEdit, QPushButton
from PyQt5.QtSerialPort import QSerialPort, QSerialPortInfo
from PyQt5.QtCore import QIODevice
from PyQt5.QtCore import Qt
from pyqtgraph import PlotWidget
import pyqtgraph as pg
import openpyxl
from openpyxl.chart import BarChart, Reference
from PyQt5.QtWidgets import QFileDialog


class MyWidget(QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi('main.ui', self)


        self.maxTopValue = 250
        self.maxRightValue = 2000

        self.output_rate = 10
        self.counter_rate = 0
        self.setWindowTitle("Коагулограф")
        self.dt_now = datetime.datetime.today()
        self.dateTimeEdit.setDateTime(self.dt_now)
        self.interferences = 0
        self.connectButton.clicked.connect(self.buttonConDis)
        self.saveButton.clicked.connect(self.save)
        self.clearButton.clicked.connect(self.onClear)
        self.importButton.clicked.connect(self.onImport)
        self.calculateButon.clicked.connect(self.calculate)
        self.serial = QSerialPort()
        self.serial.setBaudRate(9600)
        self.isConnected = False
        self.graph_data = []
        self.graph.addLegend()
        self.graph.disableAutoRange()
        self.graph.setLimits(yMin=-10, yMax=100, xMin=0, xMax=self.maxRightValue)
        self.graph.setBackground('w')
        self.pen = pg.mkPen(color=(255, 0, 0))

        self.named_data_patient = ["Дата и время", "ФИО", "№ Истории болезни",
                                   "Диагноз", "Обстоятельства", "Фибриноген", "ПТИ", "МНО", "АЧТВ", "ACT", "Д-Димер", "Доп. время(сек)"]
        self.interferences = 0
        self.data_list = []
        self.norm_data_list = []
        self.time = []
        self.now_time = 0
        self.addTime = 0

        self.ports_name_list = []
        self.ports_num_list = []
        self.strok_data = ''
        self.oldstrok_data = ''

        ports = QSerialPortInfo().availablePorts()
        for port in ports:
            full_name = port.portName() + " " + port.description()
            print(full_name)
            self.ports_name_list.append(full_name)
            self.ports_num_list.append(port.portName())

        self.ports.addItems(self.ports_name_list)
        self.serial.readyRead.connect(self.onRead)

    def buttonConDis(self):
        if self.connectButton.text() == "Начать":
            self.connectButton.setText("Остановить")
            self.onConnect()
        elif self.connectButton.text() == "Остановить":
            self.connectButton.setText("Начать")
            self.onDisconnect()

    def onConnect(self):
        print("connect")
        choose_index = self.ports_name_list.index(self.ports.currentText())
        choose_com_port = self.ports_num_list[choose_index]
        print(choose_com_port)
        self.serial.setPortName(choose_com_port)
        self.serial.open(QIODevice.ReadOnly)
        self.serial.readyRead.connect(self.onRead)

    def onDisconnect(self):
        print("disconnect")
        self.serial.close()

    def onClear(self):
        self.onDisconnect()
        self.interferences = 0
        self.oldstrok_data = ''
        self.strok_data = ''
        self.data_list = []
        self.norm_data_list = []
        self.time = []
        self.graph_data = []
        self.now_time = 0
        self.nameEdit.setText('')
        self.numEdit.setText('')
        self.diagnosisEdit.clear()
        self.conditionEdit.clear()
        self.output.clear()
        self.dt_now = datetime.datetime.today()
        self.dateTimeEdit.setDateTime(self.dt_now)
        self.graph.clear()
        self.fibrinogenEdit.setValue(0)
        self.ptiEdit.setValue(0)
        self.mnoEdit.setValue(0)
        self.actvEdit.setValue(0)
        self.actEdit.setValue(0)
        self.addTimeEdit.setValue(0)
        self.ddimerEdit.setValue(0)
        self.beforeClottingSlider.setValue(0)
        self.afterClottingSlider.setValue(0)

    def onRead(self):
        try:
            data = self.serial.readLine()
            self.strok_data = str(data)[2:-1]
            if r"\n" not in self.strok_data:
                self.oldstrok_data += self.strok_data
            else:
                if self.interferences < 10:
                    self.interferences += 1
                    self.oldstrok_data = ''
                else:
                    self.data_list.append(int(str(self.oldstrok_data + self.strok_data)[0:-4]))
                    self.oldstrok_data = ''
                    self.time.append(self.now_time)
                    self.norm_data_list.append(self.data_list[-1] / self.maxTopValue * 100)
                    self.counter_rate += 1
                    if self.counter_rate >= self.output_rate:
                        self.graph.plot(self.time, self.norm_data_list, pen=self.pen)
                        self.counter_rate = 0

                    self.now_time += 0.5
        except Exception as err:
            print("err", err)

    def save(self):
        data_patient = [self.dateTimeEdit.dateTime().toString('dd.MM.yyyy hh:mm'), self.nameEdit.text(),
                        self.numEdit.text(), self.diagnosisEdit.toPlainText(), self.conditionEdit.toPlainText(),
                        self.fibrinogenEdit.value(), self.ptiEdit.value(), self.mnoEdit.value(), self.actvEdit.value(),
                        self.actEdit.value(), self.ddimerEdit.value(), self.addTimeEdit.value()]
        wb = openpyxl.Workbook()
        wb.create_sheet(title='Первый лист', index=0)
        sheet = wb['Первый лист']
        for row in range(len(self.time)):
            cell = sheet.cell(row=row + 1, column=1)
            cell.value = self.time[row]

        for row in range(len(self.data_list)):
            cell = sheet.cell(row=row + 1, column=2)
            cell.value = self.data_list[row]

        all_name_data = self.named_data_patient + self.named_graph_data
        for row in range(len(all_name_data)):
            cell = sheet.cell(row=row + 1, column=3)
            cell.value = all_name_data[row]

        all_data = data_patient + self.graph_data
        for row in range(len(all_data)):
            cell = sheet.cell(row=row + 1, column=4)
            cell.value = all_data[row]
        try:
            filename = QFileDialog.getSaveFileName(self, "Сохранить в таблицу",
                                                   str(self.nameEdit.text().split()[0]) + '_' +
                                                   self.dateTimeEdit.dateTime().toString('dd-MM-yyyy_hh-mm'), "*.xlsx")
        except:
            filename = QFileDialog.getSaveFileName(self, "Сохранить в таблицу", '', "*.xlsx")
        print(filename)
        try:
            wb.save(filename[0])
        except:
            print('Save error')

    def onImport(self):
        self.onClear()
        filename = QFileDialog.getOpenFileName(self, "Импортировать из таблицы", '', "*.xlsx")
        wb = openpyxl.load_workbook(filename[0])
        sh_names = wb.sheetnames
        sheet = wb[sh_names[0]]
        row = 1
        while True:
            row += 1
            cell = sheet.cell(row=row, column=1)
            val = cell.value
            if val == None:
                print("end")
                break
            else:
                self.time.append(val)

                cell = sheet.cell(row=row, column=2)
                val = cell.value
                self.data_list.append(val)
        norm_data_list = [int(item / self.maxTopValue * 100) for item in self.data_list]
        self.graph.plot(self.time, norm_data_list, pen=self.pen)
        self.graph.disableAutoRange()
        self.graph.setLimits(yMin=-10, yMax=100, xMin=0, xMax=self.maxRightValue)
        try:
            data_patient = []
            for row in range(len(self.named_data_patient)):
                cell = sheet.cell(row=row + 1, column=4)
                val = cell.value
                data_patient.append(val)
            datetime_str1 = data_patient[0]
            datetime1 = datetime.datetime.strptime(datetime_str1, '%d.%m.%Y %H:%M')
            self.nameEdit.setText(data_patient[1])
            self.numEdit.setText(data_patient[2])
            self.diagnosisEdit.setPlainText(data_patient[3])
            self.conditionEdit.setPlainText(data_patient[4])
            self.fibrinogenEdit.setValue(data_patient[5])
            self.ptiEdit.setValue(data_patient[6])
            self.mnoEdit.setValue(data_patient[7])
            self.actvEdit.setValue(data_patient[8])
            self.actEdit.setValue(data_patient[9])
            self.ddimerEdit.setValue(data_patient[10])
            self.addTimeEdit.setValue(data_patient[11])
        except Exception as err:
            print(err)
        try:
            sigma = 0.25 #допуск для "плато" в долях
            period = 20
            norm_data_list = [int(item / self.maxTopValue * 100) for item in self.data_list]
            data = np.array([np.array(self.time), np.array(norm_data_list)])
            self.addTime = self.addTimeEdit.value()
            # (data[0, :] - координата времени, сек; data[1, :] - значение прибора(?), соответсвующий данному моменту времени).
            datanorm = self.contour(data, period) #перестраиваем на верхние и нижние пики. datanorm - четырёхмерный массив (NumPy):
            #(datanorm[0, :] - время верхних пиков, сек; datanorm[1, :] - значение верхних пиков; datanorm[2, :] - время нижних пиков, сек; datanorm[3, :] - значение нижних пиков).
            zeroboard = self.zeropoint(datanorm, np.min(data[1, :])) #граница ухода с начального плато (плато нулей) (одномерный массив (NumPy): [0] - координата границы, сек; [1] - индекс этой точки в datanorm).
            deltamin = self.mindeltapoint(datanorm, zeroboard[1]) # точка с минимальной шириной графика (одномерный массив (NumPy): [0] - ширина; [1] - координата, сек; [2] - индекс точки в datanorm).

            self.beforeClottingSlider.setValue(int(zeroboard[0]))
            self.afterClottingSlider.setValue(int(deltamin[1]))
            self.calculate()

        except Exception as err:
            print(err)

    def calculate(self):

        try:
            self.graph.clear()
            sigma = 0.25 #допуск для "плато" в долях
            period = 20
            norm_data_list = [int(item / self.maxTopValue * 100) for item in self.data_list]
            data = np.array([np.array(self.time), np.array(norm_data_list)])
            self.addTime = self.addTimeEdit.value()
            # (data[0, :] - координата времени, сек; data[1, :] - значение прибора(?), соответсвующий данному моменту времени).
            datanorm = self.contour(data, period) #перестраиваем на верхние и нижние пики. datanorm - четырёхмерный массив (NumPy):
            #(datanorm[0, :] - время верхних пиков, сек; datanorm[1, :] - значение верхних пиков; datanorm[2, :] - время нижних пиков, сек; datanorm[3, :] - значение нижних пиков).
            zeroboard = self.zeropoint(datanorm, np.min(data[1, :])) #граница ухода с начального плато (плато нулей) (одномерный массив (NumPy): [0] - координата границы, сек; [1] - индекс этой точки в datanorm).
            zeroboard[0] = self.beforeClottingSlider.value()

            for i in range(len(datanorm[0, :])):
                if datanorm[0, i] >= zeroboard[0]:
                    zeroboard[1] = i
                    break

            deltamin = self.mindeltapoint(datanorm, zeroboard[1]) # точка с минимальной шириной графика (одномерный массив (NumPy): [0] - ширина; [1] - координата, сек; [2] - индекс точки в datanorm).
            print(deltamin)
            deltamin[1] = self.afterClottingSlider.value()

            for i in range(len(datanorm[0, :])):
                if datanorm[0, i] >= deltamin[1]:
                    deltamin[2] = i
                    deltamin[0] = datanorm[1, i] - datanorm[3, i]
                    break

            plato = self.platopoint(datanorm, zeroboard[1], deltamin, sigma) # границы центрального плато (возле deltamin) plato - двухмерный массив (NumPy):
            #([0/1/2, 0]- координата левой/правой/трёхминутной границы, сек; [0/1/2, 1] - ширина графика в точке левой/правой/трёхминутной границы).
            #трёхминутная граница - точка, отстоящая от правой границы плато на 3 минуты.

            #рисуем график
            # любой из элементоф графика можно отключить, закомментировав его
            # self.graph.plot(data[0,:],data[1,:])



            self.graph.plot(self.time, norm_data_list, pen=self.pen)
            self.graph.plot(datanorm[0,:],datanorm[1,:], pen=pg.mkPen(color=(0, 0, 255)))
            self.graph.plot(datanorm[2, :], datanorm[3, :], name="Границы", pen=pg.mkPen(color=(0, 0, 255)))
            self.graph.plot([zeroboard[0], zeroboard[0]], [np.min(data[1, :])-10, np.max(data[1, :])+10], name=f'Время до начала свёртывания, t = {zeroboard[0]} сек', pen=pg.mkPen(color=(0, 0, 0), width=2))
            self.graph.plot([deltamin[1], deltamin[1]], [np.min(data[1, :])-10, np.max(data[1, :])+10], name=f'Время до окончания свёртывания, t = {deltamin[1]} сек', pen=pg.mkPen(color=(255, 0, 255), width=2))
            self.graph.plot([plato[0, 0], plato[0, 0]], [np.min(data[1, :])-10, np.max(data[1, :])+10], name=f'Время до начала ретракции, t = {plato[0, 0]} сек', pen=pg.mkPen(color=(255, 100, 166), width=2, style=Qt.DashLine))
            self.graph.plot([plato[1, 0], plato[1, 0]], [np.min(data[1, :])-10, np.max(data[1, :])+10], name=f'Время окончания ретракции, t = {plato[1, 0]} сек', pen=pg.mkPen(color=(255, 165, 0), width=2, style=Qt.DashLine))
            self.graph.plot([plato[2, 0], plato[2, 0]], [np.min(data[1, :])-10, np.max(data[1, :])+10], name=f'Амплитуда на 3 минуте фибринолиза = {plato[2, 1]}', pen=pg.mkPen(color=(0, 100, 100), width=2, style=Qt.DashDotDotLine))
            self.graph.showGrid(x=True, y=True)




            self.named_graph_data = ["Время до начала свёртывания, сек",
                                     "Время до окончания свёртывания, сек",
                                     "Длительность свёртывания, сек",
                                     "Время до начала ретракции, сек",
                                     "Время до окончания ретракции, сек",
                                     "Длительность ретракции, сек",
                                     "Максимальная амплитуда, ед",
                                     "Минимальная амплитуда, ед",
                                     "Амплитуда на 3 минуте фибринолиза, ед",
                                     "Скорость свёртывания, ед/мин",
                                     "Скорость нарастания фибринолиза, ед/мин",
                                     "Коэффицент ректракции, %",
                                     "Коэффицент фибринолиза, %",
                                     "Активность фибринолиза, %"]


            t_clotting = round(plato[0, 0] - zeroboard[0], 2) #Время свёртывания, сек
            v_clotting = round((datanorm[1, zeroboard[1]] - plato[0, 1]) / t_clotting * 60, 2) #Скорость свёртывания, ед/сек
            cof_retr = round((datanorm[1, zeroboard[1]] - deltamin[0]) / datanorm[1, zeroboard[1]] * 100, 2) #Коэффицент ректракции, %
            v_fibrin = round((plato[2, 1] - plato[1, 1]) / 180 * 60, 2) #Скорость нарастания фибринолиза, ед/сек
            cof_fibrin = round((plato[2, 1] - deltamin[0]) / plato[2, 1] * 100, 2) #Коэффицент фибринолиза, %
            act_fibrin = round(plato[2, 1] / (datanorm[1, zeroboard[1]] - deltamin[0]) * 100, 2)

            self.graph_data = [zeroboard[0] + self.addTime, deltamin[1] + self.addTime, t_clotting, plato[0, 0] + self.addTime, plato[1, 0],
                               plato[1, 0] - plato[0, 0], datanorm[1, zeroboard[1]], deltamin[0], plato[2, 1],
                               v_clotting, v_fibrin,
                               cof_retr, cof_fibrin, act_fibrin]


            strok_output = "Время до начала свёртывания, сек" + " " * (40 - len("Время до начала свёртывания, сек")) + str(zeroboard[0] + self.addTime) + " " * (10 - len(str(zeroboard[0] + self.addTime))) + self.secToMin(zeroboard[0] + self.addTime) + "\n"
            strok_output += "Время до окончания свёртывания, сек" + " " * (40 - len("Время до окончания свёртывания, сек")) + str(deltamin[1] + self.addTime) + " " * (10 - len(str(deltamin[1] + self.addTime))) + self.secToMin(deltamin[1] + self.addTime) + "\n"
            strok_output += "Длительность свёртывания, сек" + " " * (40 - len("Длительность свёртывания, сек")) + str(t_clotting) + " " * (10 - len(str(t_clotting))) + self.secToMin(t_clotting) + "\n"
            strok_output += "Время до начала ретракции, сек" + " " * (40 - len("Время до начала ретракции, сек")) + str(plato[0, 0] + self.addTime) + " " * (10 - len(str(plato[0, 0] + self.addTime))) + self.secToMin(plato[0, 0] + self.addTime) + "\n"
            strok_output += "Время до окончания ретракции, сек" + " " * (40 - len("Время до окончания ретракции, сек")) + str(plato[1, 0] + self.addTime) + " " * (10 - len(str(plato[1, 0] + self.addTime))) + self.secToMin(plato[1, 0] + self.addTime) + "\n"
            strok_output += "Длительность ретракции, сек" + " " * (40 - len("Длительность ретракции, сек")) + str(plato[1, 0] - plato[0, 0]) + " " * (10 - len(str(plato[1, 0] - plato[0, 0]))) + self.secToMin(plato[1, 0] - plato[0, 0]) + "\n"

            strok_output += "Максимальная амплитуда, ед" + " " * (40 - len("Максимальная амплитуда, ед")) + str(datanorm[1, zeroboard[1]]) + "\n"
            strok_output += "Минимальная амплитуда, ед" + " " * (40 - len("Минимальная амплитуда, ед")) + str(deltamin[0]) + "\n"
            strok_output += "Амплитуда на 3 минуте фибринолиза, ед" + " " * (40 - len("Амплитуда на 3 минуте фибринолиза, ед")) + str(plato[2, 1]) + "\n"

            strok_output += "Скорость свёртывания, ед/мин" + " " * (40 - len("Скорость свёртывания, ед/мин")) + str(v_clotting) + "\n"
            strok_output += "Скорость нарастания фибринолиза, ед/мин" + " " * (40 - len("Скорость нарастания фибринолиза, ед/мин")) + str(v_fibrin) + "\n"

            strok_output += "Коэффицент ректракции, %" + " " * (40 - len("Коэффицент ректракции, %")) + str(cof_retr) + "\n"
            strok_output += "Коэффицент фибринолиза, %" + " " * (40 - len("Коэффицент фибринолиза, %")) + str(cof_fibrin) + "\n"
            strok_output += "Активность фибринолиза, %" + " " * (40 - len("Активность фибринолиза, %")) + str(act_fibrin) + "\n"

            self.output.setPlainText(strok_output)
        except Exception as err:
            print(err)


    def secToMin(self, sec):
        min = int(sec // 60)
        last_sec = int(sec % 60)
        return f"({str(min)} мин {str(last_sec)} сек)"


    def contour(self, data, period):
        lendata = len(data[1, :])
        datanorm = np.zeros((4, lendata//period+1))


        maxloc = 0
        minloc = 0
        countnorm = 0

        for i in range(0, lendata, period):
            maxloc = np.argmax(data[1, i:i+period])
            minloc = np.argmin(data[1, i:i+period])
            datanorm[0,countnorm] = data[0,i+maxloc]
            datanorm[1,countnorm] = data[1,i+maxloc]
            datanorm[2,countnorm] = data[0,i+minloc]
            datanorm[3,countnorm] = data[1,i+minloc]
            countnorm = countnorm+1
        return datanorm


    def zeropoint(self, datanorm, min_of_data): #min_of_data = np.min(data[1, :])
        zeroboard = np.array([0,0])
        for countnorm in range(len(datanorm[1, :])):
            if (countnorm>2) and (datanorm[3, countnorm] > min_of_data) and (datanorm[3, countnorm-1] > min_of_data) and (datanorm[3, countnorm-2] >= min_of_data):
                zeroboard[0] = datanorm[2, countnorm-2]# + self.addTime
                zeroboard[1] = countnorm-2
                break
        return zeroboard


    def mindeltapoint(self, datanorm, zeroindex): #zeroindex = zeroboard[1]
        height = np.max(datanorm[1,:])-np.min(datanorm[1, :])
        deltamin = np.array([height, 0, 0])
        for countnorm in range(zeroindex,len(datanorm[1, :])):
            delta = datanorm[1, countnorm] - datanorm[3, countnorm]
            if (delta < deltamin[0]):
                deltamin[0] = delta
                deltamin[1] = (datanorm[0, countnorm]+datanorm[2, countnorm])/2
                deltamin[2] = countnorm
        return deltamin


    def platopoint(self, datanorm, zeroindex, deltamin, sigma):
        plato = np.array([[deltamin[1],deltamin[0]], [deltamin[1], deltamin[0]], [0,0]])
        stopper = 0
        rightindex = int(deltamin[2])
        for i in range (int(deltamin[2]), int(zeroindex), -1):
            deltanext = datanorm[1, i] - datanorm[3, i]
            delta = datanorm[1, i-1] - datanorm[3, i-1]
            deltalast = datanorm[1, i-2] - datanorm[3, i-2]
            if (delta <= deltamin[0]*(1+sigma)) and (deltanext <= deltamin[0]*(1+sigma)) and (deltalast <= deltamin[0]*(1+sigma)):
                plato[0,0] = (datanorm[0,i-2]+datanorm[2, i-2])/2# + self.addTime
                plato[0,1] = deltalast

        for i in range (int(deltamin[2]), len(datanorm[1, :])):
            deltanext = datanorm[1, i] - datanorm[3, i]
            delta = datanorm[1, i-1] - datanorm[3, i-1]
            deltalast = datanorm[1, i-2] - datanorm[3, i-2]
            if (delta <= deltamin[0]*(1+sigma)) and (deltanext <= deltamin[0]*(1+sigma)) and (deltalast <= deltamin[0]*(1+sigma)):
                plato[1,0] = (datanorm[0,i]+datanorm[2, i])/2
                plato[1,1] = deltanext
                rightindex = i

        rightindex = np.min([rightindex+18, len(datanorm[1, :])-1])
        plato[2,0] = (datanorm[0,rightindex]+datanorm[2, rightindex])/2 #+3минуты
        plato[2,1] = datanorm[1, rightindex] - datanorm[3, rightindex]
        return plato


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyWidget()
    ex.show()
    sys.exit(app.exec_())
