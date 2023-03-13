import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl import load_workbook

# Чтение данных из экселя
def readerxl():
    wb = Workbook()
    wb = load_workbook("Наумова_12-03-2023_19-13.xlsx")
    sheet = wb['Первый лист']
    lendata = len(sheet['A'])
    data = np.zeros((2, lendata ))

    for i in range(lendata):
        data[0,i] =  sheet.cell(row=i+1, column=1).value #время
        data[1,i] =  sheet.cell(row=i+1, column=2).value #значение
    return data

#нормировка
def contour(data, period):
    lendata = len(data[1, :])
    datanorm = np.zeros((4, lendata//period+1))
    
    
    maxloc = 0
    minloc = 0
    countnorm = 0
    
    
    for i in range(0, lendata, period):
        maxloc = np.argmax(data[1, i:i+period])
        minloc = np.argmin(data[1, i:i+period])
        datanorm[0,countnorm] = data[0,i+maxloc]
        datanorm[1,countnorm] = data[1,i+maxloc]
        datanorm[2,countnorm] = data[0,i+minloc]
        datanorm[3,countnorm] = data[1,i+minloc]
        countnorm = countnorm+1
    return datanorm

def zeropoint(datanorm, min_of_data): #min_of_data = np.min(data[1, :])
    zeroboard = np.array([0,0])
    for countnorm in range(len(datanorm[1, :])):

        if (countnorm>2) and (datanorm[3, countnorm] > min_of_data) and (datanorm[3, countnorm-1] > min_of_data) and (datanorm[3, countnorm-2] >= min_of_data):
            zeroboard[0] = datanorm[2, countnorm-2]
            zeroboard[1] = countnorm-2
            break
    return zeroboard


def mindeltapoint(datanorm, zeroindex): #zeroindex = zeroboard[1]
    height = np.max(datanorm[1,:])-np.min(datanorm[1, :])
    deltamin = np.array([height, 0])
    for countnorm in range(zeroindex,len(datanorm[1, :])):
        delta = datanorm[1, countnorm] - datanorm[3, countnorm]
        if (delta < deltamin[0]):
            deltamin[0] = delta
            deltamin[1] = (datanorm[0, countnorm]+datanorm[2, countnorm])/2
    return deltamin

    
def platopoint(datanorm, zeroindex, deltamin, sigma):
    plato = np.array([[0,0], [0, 0], [0,0]])
    stopper = 0
    rightindex = 0
    for i in range(zeroindex, len(datanorm[1, :])):
        deltanext = datanorm[1, i] - datanorm[3, i]
        delta = datanorm[1, i-1] - datanorm[3, i-1]
        deltalast = datanorm[1, i-2] - datanorm[3, i-2]
        if (delta <= deltamin[0]*(1+sigma)) and (deltanext <= deltamin[0]*(1+sigma)) and (deltalast <= deltamin[0]*(1+sigma)):
            if stopper == 0 :
                stopper = 1
                plato[0,0] = (datanorm[0,i-2]+datanorm[2, i-2])/2
                plato[0,1] = deltalast
            plato[1,0] = (datanorm[0,i]+datanorm[2, i])/2
            plato[1,1] = deltanext
            rightindex = i
    rightindex = np.min([rightindex+18, len(datanorm[1, :])-1])
    plato[2,0] = (datanorm[0,rightindex]+datanorm[2, rightindex])/2 #+3минуты
    plato[2,1] = datanorm[1, rightindex] - datanorm[3, rightindex]
    return plato


######

sigma = 0.25 #допуск для "плато" в долях
period = 20 

data = readerxl() #чтение данных из экселя (в основновном коде не нужен, модуль чтения уже есть). data - двухмерный массив (NumPy):
# (data[0, :] - координата времени, сек; data[1, :] - значение прибора(?), соответсвующий данному моменту времени).
datanorm = contour(data, period) #перестраиваем на верхние и нижние пики. datanorm - четырёхмерный массив (NumPy):
#(datanorm[0, :] - время верхних пиков, сек; datanorm[1, :] - значение верхних пиков; datanorm[2, :] - время нижних пиков, сек; datanorm[3, :] - значение нижних пиков).
zeroboard = zeropoint(datanorm, np.min(data[1, :])) #граница ухода с начального плато (плато нулей) (одномерный массив (NumPy): [0] - координата границы, сек; [1] - индекс этой точки в datanorm).
deltamin = mindeltapoint(datanorm, zeroboard[1]) # точка с минимальной шириной графика (одномерный массив (NumPy): [0] - ширина; [1] - координата, сек).
plato = platopoint(datanorm, zeroboard[1], deltamin, sigma) # границы центрального плато (возле deltamin) plato - двухмерный массив (NumPy): 
#([0/1/2, 0]- координата левой/правой/трёхминутной границы, сек; [0/1/2, 1] - ширина графика в точке левой/правой/трёхминутной границы).
#трёхминутная граница - точка, отстоящая от правой границы плато на 3 минуты.

#рисуем график
# любой из элементоф графика можно отключить, закомментировав его
plt.plot(data[0,:],data[1,:])
plt.plot(datanorm[0,:],datanorm[1,:], color='green')
plt.plot(datanorm[2,:],datanorm[3,:], color='red')
plt.plot([zeroboard[0], zeroboard[0]],[np.min(data[1, :])-10, np.max(data[1, :])+10],  '--',color = 'blue',label='граница нулевого плато, t = {} сек'.format(zeroboard[0]))
plt.plot([deltamin[1], deltamin[1]],[np.min(data[1, :])-10, np.max(data[1, :])+10],  '--',color = 'black',label='минимальная ширина графика, ширина = {}, t = {} сек'.format(deltamin[0], deltamin[1]))
plt.plot([plato[0,0], plato[0,0]],[np.min(data[1, :])-10, np.max(data[1, :])+10],  '--',color = 'red',label='левая граница плато, ширина = {}, t = {} сек'.format(plato[0,1], plato[0,0]))
plt.plot([plato[1,0], plato[1,0]],[np.min(data[1, :])-10, np.max(data[1, :])+10],  '--',color = 'green', label='правая граница плато, ширина = {}, t = {} сек'.format(plato[1,1], plato[1,0]))
plt.plot([plato[2,0], plato[2,0]],[np.min(data[1, :])-10, np.max(data[1, :])+10],  '--',color = 'yellow', label='3 минуты после правой границы плато, ширина = {}, t = {} сек'.format(plato[2,1], plato[2,0]))
plt.xlabel('time, sec.')

plt.grid('on')
plt.legend()
plt.show()
